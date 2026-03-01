"""
Auto-generate playstyle-weightings.json from role_weightage.xlsx
This ensures the JSON is always in sync with the Excel source of truth.

Usage: python scripts/generatePlaystyleWeightingsFromExcel.py
"""

import openpyxl
import json
import sys
from datetime import datetime

# Set UTF-8 encoding for Windows console
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

print("=" * 80)
print("Auto-Generate playstyle-weightings.json from Excel")
print("=" * 80)
print()

# Load Excel workbook
print("Loading Excel file: src/data/config/role_weightage.xlsx")
wb = openpyxl.load_workbook('src/data/config/role_weightage.xlsx')
print(f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}\n")

# Initialize output structure
output = {
    "version": "1.0.0",
    "description": "Playstyle attribute weightings for calculating PlaystyleRatings (0-100 scale)",
    "lastUpdated": datetime.now().strftime("%Y-%m-%d"),
    "notes": [
        "PlaystyleRating formula: Σ(attribute × weight) / max_possible × 100",
        "Max possible = Σ(weight × 20) for attributes on 1-20 scale",
        "All weightages are editable - adjust to change playstyle calculations",
        "Higher weight = more important for that playstyle",
        "AUTO-GENERATED from role_weightage.xlsx - DO NOT EDIT MANUALLY"
    ],
    "batting": {},
    "bowling": {
        "pace": {},
        "spin": {}
    },
    "fielding": {},
    "roleCategories": {}
}

# ============================================================================
# PROCESS BATTING SHEET (Sheet 1)
# ============================================================================
print("Processing Sheet 1: Batting Playstyles")
ws_batting = wb['Sheet1']

# Get headers
headers = [cell.value.strip() if cell.value else '' for cell in list(ws_batting.rows)[0]]
print(f"Headers: {headers}\n")

# Map Excel column names to JSON attribute names
excel_to_json = {
    'Technique': 'technique',
    'Timing': 'timing',
    'Footwork': 'footwork',
    'Placement': 'placement',
    'Range360': 'range360',
    'DefensiveShots': 'defensiveShots',
    'NeutralShots': 'neutralShots',
    'AttackingShots': 'attackingShots',
    'vsPace': 'vsPace',
    'vsSpin': 'vsSpin',
    'Judgement': 'judgement',
    'Strength': 'strength',
    'Speed': 'speed',
    'Aggression': 'aggression',
    'Creativity': 'creativity',
    'Concentration': 'concentration'
}

batting_count = 0

for row in list(ws_batting.rows)[1:-1]:  # Skip header and sum row
    archetype = row[0].value
    if not archetype or str(archetype).startswith('Sum'):
        continue

    # Clean archetype name - replace various dash types with hyphen
    archetype = str(archetype).strip().replace('�', '-').replace('–', '-').replace('—', '-')

    # Build attributes dictionary
    attributes = {}
    for i, header in enumerate(headers[1:-1], 1):  # Skip archetype name and sum
        if header.strip() in excel_to_json:
            json_key = excel_to_json[header.strip()]
            value = row[i].value
            # Convert to int if it's a number
            if isinstance(value, (int, float)):
                attributes[json_key] = int(value)

    # Add to output
    output['batting'][archetype] = {
        "description": f"Batting playstyle: {archetype}",
        "attributes": attributes
    }

    batting_count += 1
    print(f"  ✓ {archetype}: {len(attributes)} attributes")

print(f"\nProcessed {batting_count} batting playstyles\n")

# ============================================================================
# PROCESS BOWLING SHEET (Sheet 2)
# ============================================================================
print("Processing Sheet 2: Bowling Playstyles")
ws_bowling = wb['Sheet2']

# Get headers
headers_bowl = [cell.value.strip() if cell.value else '' for cell in list(ws_bowling.rows)[0]]
print(f"Headers: {headers_bowl}\n")

# Define pace and spin playstyles
pace_archetypes = {
    'Swing Bowler': 'Exploits new ball movement with swing and seam',
    'Hit-the-Deck Seamer': 'Hard length specialist, relentless pressure bowler',
    'Short-Ball Specialist': 'Intimidating pace bowler using short-pitched deliveries',
    'Death Specialist': 'Death overs yorker specialist, calm under pressure'
}

spin_archetypes = {
    'Classical Spinner': 'Traditional spinner with flight, dip, and deceptive loop',
    'Flat Spinner': 'Quick, skiddy spinner focused on accuracy and economy',
    'Mystery Spinner': 'Deceptive spinner with multiple variations and trick balls',
    'Containment Spinner': 'Metronomic line and length specialist, dot ball pressure'
}

pace_count = 0
spin_count = 0

for row in list(ws_bowling.rows)[1:]:  # Skip header row only
    if not row[0].value:
        continue

    archetype = str(row[0].value).strip()

    # Skip if empty, starts with formula, or doesn't look like an archetype
    if not archetype or archetype.startswith('=') or len(archetype) < 3:
        continue

    print(f"  Processing: '{archetype}'")  # Debug

    # Get values from Excel - skip if any value is a formula or None
    # Columns: Accuracy, Speed/Turn, Swing/Flight, Variations, Intelligence,
    #          DefensiveBowling, NeutralBowling, AttackingBowling, Stamina, Temperament
    try:
        accuracy = int(row[1].value) if isinstance(row[1].value, (int, float)) else 0
        speed_or_turn = int(row[2].value) if isinstance(row[2].value, (int, float)) else 0
        swing_or_flight = int(row[3].value) if isinstance(row[3].value, (int, float)) else 0
        variations = int(row[4].value) if isinstance(row[4].value, (int, float)) else 0
        intelligence = int(row[5].value) if isinstance(row[5].value, (int, float)) else 0
        defensive = int(row[6].value) if isinstance(row[6].value, (int, float)) else 0
        neutral = int(row[7].value) if isinstance(row[7].value, (int, float)) else 0
        attacking = int(row[8].value) if isinstance(row[8].value, (int, float)) else 0
        stamina = int(row[9].value) if isinstance(row[9].value, (int, float)) else 0
        temperament = int(row[10].value) if isinstance(row[10].value, (int, float)) else 0
    except (ValueError, TypeError):
        print(f"    [!] Skipping {archetype} - invalid data")
        continue

    # Determine if pace or spin
    if archetype in pace_archetypes:
        # Pace bowler
        attributes = {
            "accuracy": accuracy,
            "bowlingSpeed": speed_or_turn,
            "swing": swing_or_flight,
            "turn": 0,
            "variations": variations,
            "intelligence": intelligence,
            "defensiveBowling": defensive,
            "neutralBowling": neutral,
            "attackingBowling": attacking,
            "stamina": stamina,
            "temperament": temperament
        }

        output['bowling']['pace'][archetype] = {
            "description": pace_archetypes[archetype],
            "attributes": attributes
        }
        pace_count += 1
        print(f"  ✓ [PACE] {archetype}: speed={speed_or_turn}, swing={swing_or_flight}")

    elif archetype in spin_archetypes:
        # Spin bowler
        attributes = {
            "accuracy": accuracy,
            "bowlingSpeed": 0,
            "swing": 0,
            "turn": speed_or_turn,
            "flight": swing_or_flight,
            "variations": variations,
            "intelligence": intelligence,
            "defensiveBowling": defensive,
            "neutralBowling": neutral,
            "attackingBowling": attacking,
            "stamina": stamina,
            "temperament": temperament
        }

        output['bowling']['spin'][archetype] = {
            "description": spin_archetypes[archetype],
            "attributes": attributes
        }
        spin_count += 1
        print(f"  ✓ [SPIN] {archetype}: turn={speed_or_turn}, flight={swing_or_flight}")
    else:
        print(f"  [!] Unknown archetype: {archetype}")

print(f"\nProcessed {pace_count} pace and {spin_count} spin bowling playstyles\n")

# ============================================================================
# ADD FIELDING PLAYSTYLES (not in Excel - hardcoded)
# ============================================================================
print("Adding fielding playstyles (not in Excel)")

output['fielding']['Wicketkeeper'] = {
    "description": "Specialist wicketkeeper - glovework, stumping ability, and catching behind the stumps",
    "attributes": {
        "keeping": 8,
        "collecting": 5,
        "stumping": 4,
        "reflexes": 3
    }
}

print("  ✓ Wicketkeeper: keeping=8, collecting=5, stumping=4, reflexes=3\n")

# ============================================================================
# ADD ROLE CATEGORIES (defines which playstyles apply to which roles)
# ============================================================================
print("Adding role categories")

output['roleCategories'] = {
    "batsman": [
        "Opener - Slogger", "Opener - Balanced", "Opener - Anchor",
        "Top Order - Slogger", "Top Order - Balanced", "Top Order - Anchor",
        "Middle Order - Slogger", "Middle Order - Balanced", "Middle Order - Anchor",
        "Lower Order - Slogger", "Lower Order - Balanced", "Lower Order - Anchor",
        "Finisher", "Runner", "Pinch-Hitter", "Wall"
    ],
    "bowler": {
        "pace": ["Swing Bowler", "Hit-the-Deck Seamer", "Short-Ball Specialist", "Death Specialist"],
        "spin": ["Classical Spinner", "Flat Spinner", "Mystery Spinner", "Containment Spinner"]
    },
    "all-rounder": {
        "batting": [
            "Opener - Slogger", "Opener - Balanced", "Opener - Anchor",
            "Top Order - Slogger", "Top Order - Balanced", "Top Order - Anchor",
            "Middle Order - Slogger", "Middle Order - Balanced", "Middle Order - Anchor",
            "Lower Order - Slogger", "Lower Order - Balanced", "Lower Order - Anchor",
            "Finisher", "Runner", "Pinch-Hitter", "Wall"
        ],
        "bowling": {
            "pace": ["Swing Bowler", "Hit-the-Deck Seamer", "Short-Ball Specialist", "Death Specialist"],
            "spin": ["Classical Spinner", "Flat Spinner", "Mystery Spinner", "Containment Spinner"]
        }
    },
    "wicket-keeper": {
        "batting": [
            "Opener - Slogger", "Opener - Balanced", "Opener - Anchor",
            "Top Order - Slogger", "Top Order - Balanced", "Top Order - Anchor",
            "Middle Order - Slogger", "Middle Order - Balanced", "Middle Order - Anchor",
            "Lower Order - Slogger", "Lower Order - Balanced", "Lower Order - Anchor",
            "Finisher", "Runner", "Pinch-Hitter", "Wall"
        ],
        "fielding": ["Wicketkeeper"]
    }
}

print("  ✓ Role categories configured\n")

# ============================================================================
# SAVE OUTPUT
# ============================================================================
output_path = 'src/data/config/playstyle-weightings.json'
backup_path = f'src/data/config/playstyle-weightings_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'

# Create backup
try:
    with open(output_path, 'r') as f:
        existing = f.read()
    with open(backup_path, 'w') as f:
        f.write(existing)
    print(f"✓ Backup created: {backup_path}")
except FileNotFoundError:
    print("  (No existing file to backup)")

# Save new file
print(f"\nSaving to: {output_path}")
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(output, f, indent=2)

# Validate
with open(output_path, 'r', encoding='utf-8') as f:
    validate = json.load(f)

print("\n" + "=" * 80)
print("VALIDATION")
print("=" * 80)
print(f"✓ File is valid JSON")
print(f"✓ Batting playstyles: {len(validate['batting'])}")
print(f"✓ Pace bowling playstyles: {len(validate['bowling']['pace'])}")
print(f"✓ Spin bowling playstyles: {len(validate['bowling']['spin'])}")
print(f"✓ Fielding playstyles: {len(validate['fielding'])}")
print(f"✓ Role categories: {len(validate['roleCategories'])}")

print("\n" + "=" * 80)
print("SUCCESS!")
print("=" * 80)
print(f"\n✓ Generated playstyle-weightings.json from Excel")
print(f"✓ Total playstyles: {len(validate['batting']) + len(validate['bowling']['pace']) + len(validate['bowling']['spin']) + len(validate['fielding'])}")
print(f"✓ Last updated: {output['lastUpdated']}")
print(f"\nNOTE: This file is auto-generated. To make changes, edit role_weightage.xlsx")
print("      and re-run this script.\n")

# Delete backup if validation passed
import os
if os.path.exists(backup_path):
    os.remove(backup_path)
    print(f"✓ Backup deleted (validation successful)\n")
